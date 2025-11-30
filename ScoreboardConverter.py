"""

    Authors: Gregory Mejia
    Date: 11/24/2025

"""

# -- Libraries -- #

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.dimensions import SheetFormatProperties

import requests, json
from typing import cast

# -- Defaults -- #

def_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
def_font = Font(name='Aptos Narrow', size=11, bold=False, italic=False)
header_font = Font(name='Aptos Narrow', size=11, bold=True, italic=False)
thin_black_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# -- Variables -- #

image_loc = r'TopPictureCyberPatriot.png'
sb_team_scores = "https://scoreboard.uscyberpatriot.org/api/team/scores.php"

# -- Functions -- #


def populate_spreadsheet():
# def populate_spreadsheet(workbook: Workbook):
    """
        Populates a given spreadsheet with the Score Board data
    """
    scores = requests.get(sb_team_scores)
    with open("whatft.json", "w") as file:
        json.dump(scores.json(), file, indent=4)
    # print(scores.json)
    ...


def forceSpreadSheetFileExtension(string: str) -> str:
    """
        Force parameter 'string' to have .xlsx at the end of its name.
        Used to ensure a valid file type from the file extension.
    """
    try:
        dot_pos: int = string.index(".")
        if (string[dot_pos:-1] == ".xlsx"):
            return string
        else:
            return string[0:dot_pos] + "xlsx"
    except ValueError:
        return string + ".xlsx"


def createSpreadSheet(sheet_title: str, file_name: str = "untitled.xlsx"):
    """
        Creates a new spreadsheet with the given file_name.
        Styled to be like the CyberPatriot's default sheet.
    """
    # Create a new workbook and set the name to a valid version
    file_name = forceSpreadSheetFileExtension(file_name)
    workbook = Workbook()

    # Setting the main spreadsheet sheet to the main one and giving it a title
    sheet = cast(Worksheet, workbook.active)
    sheet.title = sheet_title
    sheet_prop: SheetFormatProperties = sheet.sheet_format
    sheet_prop.defaultColWidth = 8.43 + 0.72

    # Cell Merging
    sheet.merge_cells("A6:H6")  # Division Name
    sheet.merge_cells("A7:D7")  # Open Max Score
    sheet.merge_cells("E7:H7")  # Service Max Score
    sheet.merge_cells("A8:H8")  # Max Image Score
    sheet.merge_cells("A9:H9")  # Max Cisco Score

    # Headers 'n Stuff (boringg...)
    sheet["A10"] = "Team Number"
    sheet["B10"] = "Location"
    sheet["C10"] = "Division"
    sheet["D10"] = "Image Score"
    sheet["E10"] = "Adjustment"
    sheet["F10"] = "Quiz"
    sheet["G10"] = "PT"
    sheet["H10"] = "Total"

    sheet["A6"] = "HIGH SCHOOL DIVISIONS"
    sheet["A7"] = "Open Max Score = "
    sheet["E7"] = "All Service Max Score = "
    sheet["A8"] = "Network Security Score (Images) = "
    sheet["A9"] = "Cisco NetAcad Challenge = "

    # Width formatting... sighh... excel manual adjustment...
    sheet.column_dimensions["A"].width = 11.71 + 0.71
    sheet.column_dimensions["B"].width = 12.14 + 0.71
    sheet.column_dimensions["C"].width = 11.86 + 0.72
    sheet.column_dimensions["D"].width = 9.86 + 0.72
    sheet.column_dimensions["E"].width = 14.71 + 0.71

    # Why the hell is this 0.67 and the others are 0.71 or 0.72
    sheet.column_dimensions["F"].width = 8.67 + 0.67

    sheet.column_dimensions["G"].width = 8.43 + 0.71
    sheet.column_dimensions["H"].width = 8.86 + 0.72
    sheet.row_dimensions[10].height = 30

    # Style formatting :(
    for col in sheet.iter_cols():
        for cell in col:
            # Center EVERYTHING
            cell.alignment = def_align
            # From rows [6, 10]
            if (cell.row <= 10 and cell.row >= 6):
                cell.font = header_font
                if (cell.column <= 8):  # Also add those borders :D
                    cell.border = thin_black_border
            else:  # Everything else the other font
                cell.font = def_font

    # Adding the CyberPatriot image to the top of the spreadsheet
    img = Image(image_loc)
    # Divide by 1.5 because original image is 1.5x bigger when import -> Excel
    img.width /= 1.5
    img.height /= 1.5
    img.anchor = 'A1'  # Anchor it to the first cell like the original

    # Save to the file to ensure every change is reflected in a new file
    sheet.add_image(img)
    workbook.save(file_name)


# createSpreadSheet("wada", "test")
populate_spreadsheet()
print("done")
